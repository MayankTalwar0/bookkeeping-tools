from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def format_credit_policy_excel(output_file):
    """Apply formatting to the Credit Policy Excel file"""
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Define border style that actually works
    border_style = Side(border_style="thin", color="000000")
    thin_border = Border(left=border_style, right=border_style, 
                        top=border_style, bottom=border_style)
    
    # Define conditional formatting styles
    high_value_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    high_value_font = Font(color='006100')
    high_risk_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    high_risk_font = Font(color='9C0006')
    
    # Get column positions and dimensions
    header_row = next(ws.iter_rows(min_row=1, max_row=1))  # Get header row
    last_col = len(header_row)
    last_row = ws.max_row
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Header formatting
            if cell.row == 1:
                cell.font = Font(bold=True)
    
    # Apply conditional formatting to data rows
    high_value_col = None
    risk_col = None
    
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == 'High_Value':
            high_value_col = idx
        elif cell.value == 'Risk':
            risk_col = idx
    
    if high_value_col and risk_col:
        for row in ws.iter_rows(min_row=2, max_row=last_row):
            high_value_cell = row[high_value_col-1]
            risk_cell = row[risk_col-1]
            
            if high_value_cell.value == 'Yes':
                for cell in row:
                    cell.fill = high_value_fill
                    cell.font = high_value_font
            elif risk_cell.value == 'High':
                for cell in row:
                    cell.fill = high_risk_fill
                    cell.font = high_risk_font
    
    # Hide gridlines
    ws.sheet_view.showGridLines = False
    
    # Set optimal column width (just enough to fit content)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get column letter
        
        for cell in col:
            try:
                # Get length of the cell value
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Set width with minimal padding (1.2 is a conservative multiplier)
        adjusted_width = (max_length + 2) * 1.0  # Reduced padding
        ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap at 50
    
    # Save the formatted workbook
    wb.save(output_file)


def apply_credit_policy_styling(df):
    """Apply consistent styling to the credit policy DataFrame for Streamlit display"""
    
    def apply_row_styling(row):
        """Apply all styling rules to entire row based on conditions"""
        styles = [''] * len(row)
        
        # Priority 1: High Value rows (green background, dark green text)
        if 'High_Value' in row.index and row['High_Value'] == 'Yes':
            styles = ['background-color: #90EE90; color: #006400; font-weight: bold'] * len(row)
            return styles
        
        # Priority 2: High Risk rows (red background, dark red text)
        if 'Risk' in row.index and row['Risk'] == 'High':
            styles = ['background-color: #FFB6C1; color: #8B0000; font-weight: bold'] * len(row)
            return styles
        
        # Priority 3: Late Fee highlighting (for rows that aren't High Value or High Risk)
        if 'Late_Fee_Applicable' in row.index and row['Late_Fee_Applicable']:
            late_fee_index = row.index.get_loc('Late_Fee_Applicable')
            styles[late_fee_index] = 'color: red; font-weight: bold'
        
        return styles
    
    # Apply ONLY the row-level styling - no other styling methods
    styled_df = df.style.apply(apply_row_styling, axis=1)
    
    # Apply number formatting
    styled_df = styled_df.format({
        'Reduction_in_Term_Days': '{:.0f}',
        'Revised_Credit_Policy': '{:.0f}'
    })
    
    # Apply table properties
    styled_df = styled_df.set_properties(**{
        'text-align': 'center',
        'border': '1px solid black'
    })
    
    # Apply header styling
    styled_df = styled_df.set_table_styles([{
        'selector': 'th',
        'props': [('background-color', '#2c3e50'), 
                  ('color', 'white'),
                  ('font-weight', 'bold')]
    }])
    
    return styled_df